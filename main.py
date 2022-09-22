'''Script de Atualizações de Ordem de Serviço!

Funções Básicas:

1 - Consultar Ordens de Serviço

2 - Inserir contas a receber das ordens de serviço

3 - Colocar as ordens de serviço em execução

4 - criar tabela de log

5 - Caso a ordem de serviço seja cancelada ou reaberta, excluir o contas a receber.

'''

import cx_Oracle as cxo
import pandas as pd
from openpyxl import Workbook

con_orcl = cxo.connect("DTS/WDTS01@192.168.2.3/WINT")
cursor = con_orcl.cursor()

sql1 = """ SELECT NUMOS, CODCLI FROM PCORDEMSERVICO"""

cursor.execute(sql1)

result = cursor.fetchall()

df = pd.DataFrame(result, columns=['NUMOS', 'CODCLI'],)

wb = Workbook()

ws = wb.active

ws['A1'], ws['B1'] = 'NUMOS', 'CODCLI'

ws.title = 'Teste'


for ordem in result:
    ws.append(ordem)

wb.save('teste.xlsx')

print(df)








